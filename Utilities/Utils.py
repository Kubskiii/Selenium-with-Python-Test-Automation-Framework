import inspect
import logging
import softest
from openpyxl import load_workbook
import csv

class Utils(softest.TestCase):

    def AssertListItems(self,list,value):

        for stop in list:
            print("The Text is: " + stop.text)
            # assert value in stop.text
            self.soft_assert(self.assertIn,value,stop.text)
            if value in stop.text:
                print("Assert Pass")
            else:
                print("Assert Fail")
        self.assert_all()

    def CustomLogger(logLevel=logging.DEBUG):
        loggerName = inspect.stack()[1][3]

        logger = logging.getLogger(loggerName)
        logger.setLevel(logLevel)

        fileHandler = logging.FileHandler("automation.log", mode='a')

        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s - %(message)s', datefmt='%d/%b/%y %H:%M:%S %p')

        fileHandler.setFormatter(formatter)

        logger.addHandler(fileHandler)
        return logger

    def ReadDataFromExcel(fileName,sheetName):

        dataList = []
        workbook = load_workbook(filename=fileName)
        sheet = workbook[sheetName]
        rowCount = sheet.max_row
        columnCount = sheet.max_column

        for i in range(2, rowCount + 1):
            row = []
            for j in range(1, columnCount + 1):
                row.append(sheet.cell(row=i, column=j).value)
            dataList.append(row)

        return dataList

    def ReadDataFromCSV(fileName):
        dataList = []

        csvData = open(fileName,"r")
        reader = csv.reader(csvData)

        next(reader)
        for row in reader:
            dataList.append(row)

        return dataList

    def ConvertTextToNumber(text):

        ConvertedText = []

        for element in text:
            var1 = int(element.text.replace("₹", "").replace(",", ""))
            ConvertedText.append(var1)

        return ConvertedText

    def SortLowToHigh(list):
        return sorted(list)

    def AssertTwoLists(self,list1,list2):

        for originalPrice,sortedPrice in zip(list1,list2):
            self.soft_assert(self.assertEqual,originalPrice,sortedPrice)
            if originalPrice == sortedPrice:
                print(f"Assert Pass: Expected Price {sortedPrice}, Actual Price {originalPrice}")
            else:
                print(f"Assert Fail: Expected Price {sortedPrice}, Actual Price {originalPrice}")




